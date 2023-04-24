import openpyxl

inventory_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inventory_file["Sheet1"]

products_per_supplier = {}
total_value_of_supplier = {}
products_under_10_inv = {}

for product_row in range(2,product_list.max_row + 1):
    #getting the collumns
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_number = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)
    
    #calculation of number of products per supplier
    if supplier_name in products_per_supplier:
        curr_num_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = curr_num_products + 1
    else:
        products_per_supplier[supplier_name] = 1
    
    #calculation of total value of products per supplier
    if supplier_name in total_value_of_supplier:
        curr_total_value = total_value_of_supplier.get(supplier_name)
        total_value_of_supplier[supplier_name] = curr_total_value + inventory * price
    else:
        total_value_of_supplier[supplier_name] = inventory * price
    
    #calculation of products with inventory under 10
    if inventory < 10:
        products_under_10_inv[int(product_number)] = int(inventory)
    
    #calculation of inventory price
    inventory_price.value = inventory * price
     
inventory_file.save("inventory_with_total_value.xlsx")   
print(products_per_supplier)
print(total_value_of_supplier)
print(products_under_10_inv)